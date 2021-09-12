import json
from openpyxl import Workbook, load_workbook


class dz():

    def __init__(self, _list=[]):
        self._list = _list

    def convertor(self, list1, list2):
        dictionary = dict(zip(list1, list2))
        return dictionary

    def load_json(self, dict_data):
        json.dump(dict_data, open("file.json", 'r'))

    def read_file(self, pathname):
        rw = open(pathname, "r")
        return json.load(rw)

    def exel(self, _dict):
        workbook = Workbook()
        workbook_active = workbook.active

        for _, (key, value) in enumerate(_dict.items(), start=1):
            workbook_active[key] = value

        workbook.save("output.xlsx")

    def readxl(self):
        workbook = load_workbook(filename="output.xlsx")
        workbook_active = workbook.active
        _dict = {}
        for row in list(workbook_active.rows):
            _dict[row[0].coordinate] = row[0].value
        return _dict


dzz = dz()
dzz.exel({"A1": "aaaaaaaa", "A2": "bbbbbbbb"})
print(dzz.readxl())
